# File: file_io/excel_loader.py
#
# ExcelLoader is the concrete adapter that implements ParserInterface
# using openpyxl to read/write the Money Manager Excel template.

from __future__ import annotations

from typing import Dict, List, Optional
from pathlib import Path
from datetime import datetime
import shutil

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from file_io.parser_interface import ParserInterface
from models.transaction import Transaction
from utils.general_helper import GeneralHelper

# ---- Validators (user-facing, fail-fast checks) ----
from utils.validator import (
    ValidationError,
    validate_open_excel_path,
    validate_year_value,
    validate_income_value,
    validate_month_name,
    validate_day_in_month,
    validate_slot_index,
    validate_transaction_or_none,
    validate_backup_directory,
    validate_import_into_session,
    validate_fresh_session_paths,
    validate_template_signature,
    validate_download_target_path,
    # Phase 2 additions:
    validate_date_matches_month,
    validate_upload_filesize,
)


class ExcelLoader(ParserInterface):
    """
    Abstraction Function:
    - Bridge between the Excel template and in-memory Transaction objects.
    - Knows fixed cell locations and grid layout for months/days.

    Representation Invariant:
    - _wb and _ws are non-None only when a workbook is open.
    - Template signature must match TEMPLATE_MARKER_VALUE at TEMPLATE_MARKER_CELL.
    """

    # -------- Template signature / static anchors --------
    TEMPLATE_MARKER_CELL: str = "A1"
    TEMPLATE_MARKER_VALUE: str = "MM_TMU"

    CURRENT_INCOME_CELL: str = "C5"
    YEAR_CELL: str = "B6"

    # Month anchor rows (row where the month name string appears)
    MONTH_ANCHOR_ROW: Dict[str, int] = {
        "JANUARY": 8, "FEBRUARY": 72, "MARCH": 132, "APRIL": 196,
        "MAY": 258, "JUNE": 322, "JULY": 384, "AUGUST": 448,
        "SEPTEMBER": 512, "OCTOBER": 574, "NOVEMBER": 638, "DECEMBER": 700,
    }

    # Column mapping (1-based indices)
    COL: Dict[str, int] = {
        "DATE": 1, "DAY": 3, "CATEGORY": 4, "AMOUNT": 9, "TYPE": 10, "DESCRIPTION": 11,
    }

    VALID_TYPES: set[str] = {"EXPENSE", "INCOME", "NONE"}

    def __init__(self) -> None:
        self._wb: Optional[Workbook] = None
        self._ws: Optional[Worksheet] = None
        self._path: Optional[Path] = None

    # --------------- Lifecycle ---------------

    def open(self, file_path: str) -> None:
        """
        REQUIRES: file_path points to readable .xlsx
        MODIFIES: self
        EFFECTS:  Opens the workbook and stores handles.
        """
        validate_open_excel_path(file_path)
        self._path = Path(file_path).resolve()
        self._wb = load_workbook(filename=str(self._path))
        self._ws = self._wb.active

    def close(self) -> None:
        """
        REQUIRES: none
        MODIFIES: self
        EFFECTS:  Clears workbook handles (no file writes).
        """
        self._wb = None
        self._ws = None
        self._path = None

    # --------------- Verification ---------------

    def verify_template(self) -> bool:
        """
        REQUIRES: workbook is open
        EFFECTS:  Returns True if template marker matches.
        """
        self.ensure_open()
        value = self._ws[self.TEMPLATE_MARKER_CELL].value  # type: ignore[index]
        return str(value).strip().upper() == self.TEMPLATE_MARKER_VALUE

    # --------------- Fixed fields ---------------

    def get_year(self) -> int:
        self.ensure_open()
        raw = self._ws[self.YEAR_CELL].value  # type: ignore[index]
        return int(float(raw))

    def set_year(self, year: int) -> None:
        validate_year_value(year)
        self.ensure_open()
        self._ws[self.YEAR_CELL] = int(year)  # type: ignore[index]

    def get_current_income(self) -> float:
        self.ensure_open()
        raw = self._ws[self.CURRENT_INCOME_CELL].value  # type: ignore[index]
        return self.to_float(raw)

    def set_current_income(self, income: float) -> None:
        validate_income_value(income)
        self.ensure_open()
        self._ws[self.CURRENT_INCOME_CELL] = float(income)  # type: ignore[index]

    # --------------- Month/Day I/O ---------------

    def read_month(self, month_name: str) -> Dict[int, List[Optional[Transaction]]]:
        """
        REQUIRES: month_name is a valid template month
        EFFECTS:  Returns {day -> [slot0, slot1]} for the month.
        """
        self.ensure_open()
        month_key = validate_month_name(month_name)  # normalize + validate
        anchor = self.get_anchor_row(month_key)
        year = self.get_year()
        day_count = GeneralHelper.get_days_in_month(month_key, year)

        result: Dict[int, List[Optional[Transaction]]] = {}
        for d in range(1, day_count + 1):
            row_top = self.row_for_day(anchor, d, slot_index=0)
            row_bottom = row_top + 1
            slot0 = self.read_row_as_transaction(month_key, year, d, row_top)
            slot1 = self.read_row_as_transaction(month_key, year, d, row_bottom)
            result[d] = [slot0, slot1]
        return result

    def write_day_entry(
        self, month_name: str, day: int, slot_index: int, tx: Optional[Transaction]
    ) -> None:
        """
        REQUIRES: inputs validated by validators
        MODIFIES: workbook
        EFFECTS:  Writes or clears a specific day/slot row.
        """
        self.ensure_open()
        month_key = validate_month_name(month_name)
        validate_day_in_month(month_key, self.get_year(), day)
        validate_slot_index(slot_index)
        validate_transaction_or_none(tx)

        # Phase 2: ensure the tx.date belongs to the section being written
        if tx is not None:
            validate_date_matches_month(month_key, tx.date, self.get_year())

        anchor = self.get_anchor_row(month_key)
        row = self.row_for_day(anchor, day, slot_index)

        if tx is None:
            self.clear_row(row)
            self._ws.cell(row=row, column=self.COL["TYPE"]).value = "NONE"
            return

        self._ws.cell(row=row, column=self.COL["DATE"]).value = self.extract_day(tx.date)
        self._ws.cell(row=row, column=self.COL["DAY"]).value = tx.day
        self._ws.cell(row=row, column=self.COL["CATEGORY"]).value = tx.category
        self._ws.cell(row=row, column=self.COL["AMOUNT"]).value = float(tx.amount)
        self._ws.cell(row=row, column=self.COL["TYPE"]).value = tx.type.upper()
        self._ws.cell(row=row, column=self.COL["DESCRIPTION"]).value = tx.description

    def read_all(self) -> Dict[str, Dict[int, List[Optional[Transaction]]]]:
        """
        REQUIRES: workbook is open
        EFFECTS:  Returns { month -> { day -> [slot0, slot1] } } for the whole year.
        """
        self.ensure_open()
        data: Dict[str, Dict[int, List[Optional[Transaction]]]] = {}
        for month_key in self.months_in_order():
            data[month_key] = self.read_month(month_key)
        return data

    def write_all(self, data: Dict[str, Dict[int, List[Optional[Transaction]]]]) -> None:
        """
        REQUIRES: data follows the schema
        MODIFIES: workbook
        EFFECTS:  Writes all months/days/slots back to the file buffer.
        """
        self.ensure_open()
        for month_key, days in data.items():
            month_norm = validate_month_name(month_key)
            for day, slots in days.items():
                for idx, tx in enumerate(slots):
                    self.write_day_entry(month_norm, day, idx, tx)

    # --------------- Persistence ---------------

    def save(self) -> None:
        """
        REQUIRES: workbook is open and bound to a path
        MODIFIES: workbook file
        EFFECTS:  Saves the current workbook to the same path.
        """
        self.ensure_open()
        if not self._path:
            raise ValidationError("No backing path associated with this workbook.")
        self._wb.save(str(self._path))  # type: ignore[union-attr]

    def save_as(self, file_path: str) -> None:
        """
        REQUIRES: valid target path with .xlsx
        MODIFIES: filesystem
        EFFECTS:  Saves the workbook to a new path.
        """
        validate_download_target_path(file_path)
        self.ensure_open()
        self._wb.save(file_path)  # type: ignore[union-attr]

    # --------------- Session helpers ---------------

    def get_path(self) -> Optional[str]:
        """EFFECTS: Returns absolute path of the open workbook, else None."""
        return None if self._path is None else str(self._path)

    def save_backup(self, backup_dir: str) -> str:
        """
        REQUIRES: workbook is open; backup_dir writable
        MODIFIES: filesystem
        EFFECTS:  Writes a timestamped copy of the open workbook into backup_dir.
        """
        self.ensure_open()
        if self._path is None:
            raise ValidationError("No backing path associated with this workbook.")
        validate_backup_directory(backup_dir)

        dst_dir = Path(backup_dir).resolve()
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"{self._path.stem}_backup_{ts}.xlsx"
        backup_path = dst_dir / backup_name

        self.save()  # flush current buffer first
        shutil.copy2(str(self._path), str(backup_path))
        return str(backup_path)

    def prepare_fresh_session(self, template_path: str, storage_dir: str) -> str:
        """
        REQUIRES: template_path exists; storage_dir writable
        MODIFIES: filesystem
        EFFECTS:  Copy template -> storage_dir/user_data.xlsx, open it, return its path.
        """
        validate_fresh_session_paths(template_path, storage_dir)

        src = Path(template_path).resolve()
        storage = Path(storage_dir).resolve()
        dst = storage / "user_data.xlsx"

        shutil.copy2(str(src), str(dst))

        self.open(str(dst))

        if not self.verify_template():
            self.close()
            try:
                dst.unlink(missing_ok=True)  # type: ignore[arg-type]
            except Exception:
                pass
            raise ValidationError("Copied session file failed template verification.")

        return str(dst)

    def import_into_session(self, source_path: str) -> None:
        """
        REQUIRES: workbook is open; source_path points to valid template
        MODIFIES: workbook file
        EFFECTS:  Overwrites the session file with source_path and re-opens it.
        """
        if self._path is None:
            raise ValidationError("No session file is open. Start a fresh session first.")

        # Phase 2: size guard before we touch the file
        validate_upload_filesize(source_path)

        validate_import_into_session(self, source_path)
        src = Path(source_path).resolve()

        self.save()
        shutil.copy2(str(src), str(self._path))
        self.open(str(self._path))

    # --------------- Internals (renamed to Java-like semantics) ---------------

    def ensure_open(self) -> None:
        """
        REQUIRES: workbook should be open before any I/O call
        EFFECTS:  Raises ValidationError if not open.
        """
        if self._wb is None or self._ws is None:
            raise ValidationError("Workbook is not open. Call open(file_path) first.")

    def normalize_month(self, month_name: str) -> str:
        """
        REQUIRES: month_name is non-empty
        EFFECTS:  Returns the normalized uppercase key, or raises if unknown.
        NOTE: Public callers should prefer validator.validate_month_name().
        """
        key = month_name.strip().upper()
        if key not in self.MONTH_ANCHOR_ROW:
            raise ValidationError(f"Unknown month: {month_name}")
        return key

    def get_anchor_row(self, month_key: str) -> int:
        """EFFECTS: Returns the anchor row for the month header."""
        return self.MONTH_ANCHOR_ROW[month_key]

    def row_for_day(self, anchor_row: int, day: int, slot_index: int) -> int:
        """
        REQUIRES: day >= 1; slot_index in {0,1}
        EFFECTS:  Computes the worksheet row for the given (day, slot).
        """
        if day < 1:
            raise ValidationError("day must be >= 1")
        if slot_index not in (0, 1):
            raise ValidationError("slot_index must be 0 or 1")
        return anchor_row + 2 + (day - 1) * 2 + slot_index

    def months_in_order(self) -> List[str]:
        """EFFECTS: Returns the months in template order."""
        return [
            "JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE",
            "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER"
        ]

    def clear_row(self, row: int) -> None:
        """MODIFIES: worksheet; EFFECTS: Clears all mapped columns in the row."""
        for col in self.COL.values():
            self._ws.cell(row=row, column=col).value = None  # type: ignore[union-attr]

    def read_row_as_transaction(
        self, month_key: str, year: int, day: int, row: int
    ) -> Optional[Transaction]:
        """
        REQUIRES: worksheet row corresponds to a valid slot
        EFFECTS:  Parses the row into a Transaction or returns None for blanks/NONE.
        """
        t = self.safe_cell_str(row, self.COL["TYPE"]).strip().upper()
        date_cell = self.safe_cell_str(row, self.COL["DATE"]).strip()
        day_name = self.safe_cell_str(row, self.COL["DAY"]).strip()
        category = self.safe_cell_str(row, self.COL["CATEGORY"]).strip()
        amount_raw = self._ws.cell(row=row, column=self.COL["AMOUNT"]).value  # type: ignore[union-attr]
        desc = self.safe_cell_str(row, self.COL["DESCRIPTION"]).strip()

        is_all_blank = (
            t == "" and date_cell == "" and day_name == "" and category == "" and
            desc == "" and (amount_raw is None or str(amount_raw).strip() == "")
        )
        if t == "NONE" or (t == "" and is_all_blank):
            return None
        if t not in self.VALID_TYPES:
            return None

        day_number = self.to_int(date_cell) if date_cell else day
        date_str = self.date_to_string(year, month_key, day_number)
        amount_val = self.to_float(amount_raw)

        return Transaction(
            date=date_str, day=day_name, category=category,
            amount=amount_val, type=t, description=desc,
        )

    def safe_cell_str(self, row: int, col: int) -> str:
        """EFFECTS: Returns the cell value as a string, or '' if None."""
        val = self._ws.cell(row=row, column=col).value  # type: ignore[union-attr]
        return "" if val is None else str(val)

    def to_float(self, raw) -> float:
        """EFFECTS: Best-effort float coercion for Excel values (currency-safe)."""
        if raw is None:
            return 0.0
        s = str(raw).strip()
        if s == "":
            return 0.0
        s = s.replace("$", "").replace(",", "")
        try:
            return float(s)
        except ValueError:
            return 0.0

    def to_int(self, raw) -> int:
        """EFFECTS: Best-effort int coercion for Excel values."""
        if raw is None or str(raw).strip() == "":
            return 0
        try:
            return int(float(str(raw)))
        except ValueError:
            return 0

    def date_to_string(self, year: int, month_key: str, day_number: int) -> str:
        """EFFECTS: Returns 'YYYY-MM-DD' using template month key."""
        month_num = self.month_to_number(month_key)
        return f"{year:04d}-{month_num:02d}-{day_number:02d}"

    def month_to_number(self, month_key: str) -> int:
        """EFFECTS: Maps template month key to 1–12."""
        mapping = {
            "JANUARY": 1, "FEBRUARY": 2, "MARCH": 3, "APRIL": 4,
            "MAY": 5, "JUNE": 6, "JULY": 7, "AUGUST": 8,
            "SEPTEMBER": 9, "OCTOBER": 10, "NOVEMBER": 11, "DECEMBER": 12
        }
        return mapping[month_key]

    def extract_day(self, date_str: str) -> int:
        """EFFECTS: Parses 'YYYY-MM-DD' and returns DD as int; 0 on failure."""
        try:
            parts = str(date_str).split("-")
            return int(parts[-1])
        except Exception:
            return 0
