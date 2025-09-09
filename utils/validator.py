# File: utils/validator.py
#
# Validation helpers for UI + I/O actions.
#
# Abstraction:
# - Each validate_* function raises ValidationError with a clean, user-facing
#   message when the preconditions for an action are not met.
#
# NOTE: Validators *do not* mutate files; they only check assumptions.

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional
import os
import tempfile
import datetime

# Use openpyxl directly for template signature checks to avoid circular imports
from openpyxl import load_workbook

from models.transaction import Transaction
from utils.general_helper import GeneralHelper


# --------------------------- Error Type ---------------------------

@dataclass
class ValidationError(Exception):
    """Simple, user-facing validation error."""
    message: str

    def __str__(self) -> str:
        return self.message


# --------------------------- Constants ---------------------------

_TEMPLATE_MARKER_CELL = "A1"
_TEMPLATE_MARKER_VALUE = "MM_TMU"
_MONTH_KEYS = {
    "JANUARY","FEBRUARY","MARCH","APRIL","MAY","JUNE",
    "JULY","AUGUST","SEPTEMBER","OCTOBER","NOVEMBER","DECEMBER"
}

# Local month -> number map (kept here to avoid coupling)
_MONTH_TO_NUM = {
    "JANUARY": 1, "FEBRUARY": 2, "MARCH": 3, "APRIL": 4,
    "MAY": 5, "JUNE": 6, "JULY": 7, "AUGUST": 8,
    "SEPTEMBER": 9, "OCTOBER": 10, "NOVEMBER": 11, "DECEMBER": 12,
}


# ================================================================
# === WelcomeScreen validators ===================================
# ================================================================

def validate_fresh_session_paths(template_path: str, storage_dir: str) -> None:
    """Preconditions to start a fresh session: template exists + storage dir writable."""
    tpath = Path(template_path)
    sdir = Path(storage_dir)

    if not tpath.exists() or not tpath.is_file():
        raise ValidationError(f"Template not found at: {tpath}")
    if tpath.suffix.lower() != ".xlsx":
        raise ValidationError("Template must be an .xlsx file.")

    ensure_directory_writable(sdir)


def validate_user_upload_path(file_path: str) -> None:
    """Preconditions for an uploaded Excel file: exists, .xlsx, readable."""
    if not file_path or str(file_path).strip() == "":
        raise ValidationError("Please select an Excel file (.xlsx).")

    fpath = Path(file_path)
    if not fpath.exists() or not fpath.is_file():
        raise ValidationError(f"File not found at: {fpath}")
    if fpath.suffix.lower() != ".xlsx":
        raise ValidationError("Selected file must have a .xlsx extension.")
    if not os.access(str(fpath), os.R_OK):
        raise ValidationError(f"You do not have permission to read this file:\n{fpath}")


def validate_template_signature(file_path: str) -> None:
    """Opens the workbook and checks the Money Manager template marker."""
    try:
        wb = load_workbook(filename=str(file_path))
        ws = wb.active
        marker = ws[_TEMPLATE_MARKER_CELL].value  # type: ignore[index]
        ok = str(marker or "").strip().upper() == _TEMPLATE_MARKER_VALUE
        wb.save(str(file_path))  # ensure handles are closed
    except Exception as e:
        raise ValidationError(f"Could not open the selected file.\n\n{e}")

    if not ok:
        raise ValidationError(
            "This file doesn't match the Money Manager template "
            f"(missing '{_TEMPLATE_MARKER_VALUE}' marker in cell {_TEMPLATE_MARKER_CELL})."
        )


def validate_download_target_path(save_path: str) -> None:
    """Preconditions for saving to a path: non-empty, dir writable, .xlsx extension."""
    if not save_path or str(save_path).strip() == "":
        raise ValidationError("Please choose a location to save the file.")

    target = Path(save_path)
    if target.suffix.lower() != ".xlsx":
        raise ValidationError("The target file name must end with .xlsx")

    ensure_directory_writable(target.parent)


def validate_export_target_path(target_path: str, active_session_path: Optional[str]) -> None:
    """
    REQUIRES: target_path is a string; active_session_path may be None
    EFFECTS:  Validates target is writable .xlsx AND not in live storage directory.
    """
    validate_download_target_path(target_path)
    if not active_session_path:
        return

    try:
        session_file = Path(active_session_path).resolve()
        storage_directory = session_file.parent
        target = Path(target_path).resolve()
        try:
            target.relative_to(storage_directory)
            raise ValidationError(
                "Please do not save into the application's live storage folder.\n"
                "Choose a different location outside template/file_storage."
            )
        except ValueError:
            pass
    except Exception:
        return


# ================================================================
# === ExcelLoader validators =====================================
# ================================================================

def validate_open_excel_path(file_path: str) -> None:
    """Preconditions for opening: non-empty, exists, .xlsx, readable."""
    if not file_path or str(file_path).strip() == "":
        raise ValidationError("Please provide a workbook path to open.")
    fpath = Path(file_path)
    if not fpath.exists() or not fpath.is_file():
        raise ValidationError(f"Workbook not found at: {fpath}")
    if fpath.suffix.lower() != ".xlsx":
        raise ValidationError("Workbook must be an .xlsx file.")
    if not os.access(str(fpath), os.R_OK):
        raise ValidationError(f"You do not have permission to read this workbook:\n{fpath}")


def validate_workbook_is_open(loader) -> None:
    """Confirms a loader has an open workbook (duck-typed: has get_path())."""
    if loader is None or not hasattr(loader, "get_path") or loader.get_path() is None:
        raise ValidationError("No workbook is open. Start a session or open a file first.")


def validate_year_value(year: int) -> None:
    try:
        y = int(year)
    except Exception:
        raise ValidationError("Year must be a 4-digit number (e.g., 2025).")
    if y < 1900 or y > 2100:
        raise ValidationError("Year must be between 1900 and 2100.")


def validate_income_value(amount: float) -> None:
    try:
        a = float(amount)
    except Exception:
        raise ValidationError("Income must be a number.")
    if a < 0:
        raise ValidationError("Income must be non-negative.")


def validate_month_name(month_name: str) -> str:
    """Normalize + validate month name against template months."""
    if not month_name or str(month_name).strip() == "":
        raise ValidationError("Month name is required.")
    key = str(month_name).strip().upper()
    if key not in _MONTH_KEYS:
        raise ValidationError("Unknown month name. Use January–December.")
    return key


def validate_day_in_month(month_key: str, year: int, day: int) -> None:
    """Ensure day is valid for given month/year."""
    try:
        max_day = GeneralHelper.get_days_in_month(month_key, int(year))
    except Exception:
        raise ValidationError("Could not validate the day for the given month/year.")
    try:
        d = int(day)
    except Exception:
        raise ValidationError("Day must be a whole number.")
    if d < 1 or d > max_day:
        raise ValidationError(f"Day {d} is not valid for {month_key} {year} (max = {max_day}).")


def validate_slot_index(slot_index: int) -> None:
    try:
        idx = int(slot_index)
    except Exception:
        raise ValidationError("Slot index must be 0 or 1.")
    if idx not in (0, 1):
        raise ValidationError("Slot index must be 0 or 1.")


def validate_transaction_or_none(tx: Optional[Transaction]) -> None:
    """Validate optional Transaction object (fields + type/amount/category/date)."""
    if tx is None:
        return

    try:
        amt = float(tx.amount)
    except Exception:
        raise ValidationError("Transaction amount must be a number.")
    if amt < 0:
        raise ValidationError("Transaction amount must be non-negative.")

    t = str(getattr(tx, "type", "")).strip().upper()
    if t not in {"INCOME", "EXPENSE"}:
        raise ValidationError("Transaction type must be INCOME or EXPENSE.")

    cat = str(getattr(tx, "category", "")).strip()
    if cat == "":
        raise ValidationError("Transaction category cannot be empty.")

    date_s = str(getattr(tx, "date", "")).strip()
    if date_s == "":
        raise ValidationError("Transaction date is required (YYYY-MM-DD).")
    try:
        datetime.datetime.strptime(date_s, "%Y-%m-%d")
    except Exception:
        raise ValidationError("Invalid date format. Expected YYYY-MM-DD.")


def validate_backup_directory(backup_dir: str) -> None:
    if not backup_dir or str(backup_dir).strip() == "":
        raise ValidationError("Please choose a backup folder.")
    ensure_directory_writable(Path(backup_dir))


def validate_import_into_session(loader, source_path: str) -> None:
    """Preconditions before replacing the session file with an uploaded workbook."""
    validate_workbook_is_open(loader)
    validate_user_upload_path(source_path)
    validate_template_signature(source_path)


# ================================================================
# === Extra Phase 2 validators ===================================
# ================================================================

def validate_date_matches_month(month_key: str, date_str: str, year: int) -> None:
    """Ensures tx.date’s year/month matches the target month section."""
    try:
        d = datetime.datetime.strptime(date_str, "%Y-%m-%d")
    except Exception:
        raise ValidationError("Invalid transaction date. Expected YYYY-MM-DD.")
    month_num = _MONTH_TO_NUM.get(str(month_key).strip().upper())
    if month_num is None:
        raise ValidationError("Unknown month name. Use January–December.")
    if d.year != int(year) or d.month != month_num:
        raise ValidationError(f"Date {date_str} does not belong to {month_key} {year}.")


def validate_upload_filesize(file_path: str, max_mb: int = 20) -> None:
    """Reject absurdly large uploads (> max_mb)."""
    fpath = Path(file_path)
    size_mb = fpath.stat().st_size / (1024 * 1024)
    if size_mb > max_mb:
        raise ValidationError(f"Upload too large ({size_mb:.1f} MB). Limit is {max_mb} MB.")


def validate_active_session(session_path: Optional[str]) -> None:
    """Friendly error if actions are attempted without a live session path."""
    if not session_path or not Path(session_path).exists():
        raise ValidationError("No active session found. Start a fresh or imported session first.")


# --------------------------- Helpers -----------------------------

def ensure_directory_writable(directory: Path) -> None:
    try:
        directory.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        raise ValidationError(f"Cannot create directory:\n{directory}\n\n{e}")
    try:
        with tempfile.TemporaryFile(dir=str(directory)):
            pass
    except Exception:
        raise ValidationError(f"This folder is not writable:\n{directory}")
